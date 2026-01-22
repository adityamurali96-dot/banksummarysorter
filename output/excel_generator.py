"""
Excel output generator for categorized bank transactions.

Creates a formatted Excel workbook with multiple sheets:
1. All Transactions
2. Category Summary
3. Monthly Summary
4. Flagged for Review
5. Statistics
"""
from collections import defaultdict
from datetime import date
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from parsers.base_parser import Transaction


# Style definitions
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FLAGGED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CURRENCY_FORMAT = '#,##0.00'
DATE_FORMAT = 'DD-MMM-YYYY'
PERCENT_FORMAT = '0.00%'


def generate_output_excel(
    transactions: List[Transaction],
    output_path: str,
    include_raw_text: bool = False
) -> str:
    """
    Generate an Excel workbook with categorized transactions.

    Args:
        transactions: List of categorized transactions
        output_path: Path to save the Excel file
        include_raw_text: Whether to include raw text column

    Returns:
        Path to the generated file
    """
    print(f"\nGenerating Excel output: {output_path}")

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Create sheets
    _create_all_transactions_sheet(wb, transactions, include_raw_text)
    _create_category_summary_sheet(wb, transactions)
    _create_monthly_summary_sheet(wb, transactions)
    _create_flagged_sheet(wb, transactions)
    _create_statistics_sheet(wb, transactions)

    # Save workbook
    wb.save(output_path)
    print(f"Excel file saved: {output_path}")

    return output_path


def _create_all_transactions_sheet(
    wb: Workbook,
    transactions: List[Transaction],
    include_raw_text: bool
) -> None:
    """Create the All Transactions sheet."""
    ws = wb.create_sheet("All Transactions")

    # Headers
    headers = [
        "Date", "Description", "Debit", "Credit", "Balance",
        "Category", "Subcategory", "Confidence", "Source", "Notes"
    ]
    if include_raw_text:
        headers.append("Raw Text")

    # Write header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    # Write data rows
    for row_idx, txn in enumerate(transactions, 2):
        # Date
        cell = ws.cell(row=row_idx, column=1, value=txn.date)
        cell.number_format = DATE_FORMAT

        # Description
        ws.cell(row=row_idx, column=2, value=txn.description)

        # Debit
        cell = ws.cell(row=row_idx, column=3, value=txn.debit)
        if txn.debit:
            cell.number_format = CURRENCY_FORMAT

        # Credit
        cell = ws.cell(row=row_idx, column=4, value=txn.credit)
        if txn.credit:
            cell.number_format = CURRENCY_FORMAT

        # Balance
        cell = ws.cell(row=row_idx, column=5, value=txn.balance)
        if txn.balance:
            cell.number_format = CURRENCY_FORMAT

        # Category
        ws.cell(row=row_idx, column=6, value=txn.category)

        # Subcategory
        ws.cell(row=row_idx, column=7, value=txn.subcategory)

        # Confidence
        cell = ws.cell(row=row_idx, column=8, value=txn.categorization_confidence)
        cell.number_format = PERCENT_FORMAT

        # Source
        ws.cell(row=row_idx, column=9, value=txn.categorization_source)

        # Notes (Haiku suggestion for flagged items)
        notes = txn.haiku_suggestion if txn.categorization_source == "flagged" else ""
        ws.cell(row=row_idx, column=10, value=notes)

        # Raw text (optional)
        if include_raw_text:
            ws.cell(row=row_idx, column=11, value=txn.raw_text)

        # Highlight flagged rows
        if txn.categorization_source == "flagged":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = FLAGGED_FILL

        # Alternate row colors (for non-flagged)
        elif row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = ALT_ROW_FILL

    # Set column widths
    column_widths = [12, 50, 15, 15, 15, 20, 25, 12, 10, 40]
    if include_raw_text:
        column_widths.append(60)
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[_get_column_letter(col)].width = width

    # Add autofilter
    ws.auto_filter.ref = f"A1:{_get_column_letter(len(headers))}{len(transactions) + 1}"

    # Freeze header row
    ws.freeze_panes = "A2"


def _create_category_summary_sheet(wb: Workbook, transactions: List[Transaction]) -> None:
    """Create the Category Summary sheet."""
    ws = wb.create_sheet("Category Summary")

    # Aggregate data by category and subcategory
    summary: Dict[Tuple[str, str], Dict[str, float]] = defaultdict(
        lambda: {'debit': 0.0, 'credit': 0.0, 'count': 0}
    )

    for txn in transactions:
        key = (txn.category, txn.subcategory)
        summary[key]['debit'] += txn.debit or 0
        summary[key]['credit'] += txn.credit or 0
        summary[key]['count'] += 1

    # Headers
    headers = ["Category", "Subcategory", "Total Debit", "Total Credit", "Net", "Count"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Sort by category, then subcategory
    sorted_keys = sorted(summary.keys())

    row_idx = 2
    current_category = None
    category_totals: Dict[str, Dict[str, float]] = defaultdict(
        lambda: {'debit': 0.0, 'credit': 0.0, 'count': 0}
    )

    for category, subcategory in sorted_keys:
        data = summary[(category, subcategory)]

        # Category subtotal row (when category changes)
        if current_category and current_category != category:
            _write_category_subtotal(ws, row_idx, current_category, category_totals[current_category])
            row_idx += 1

        current_category = category
        category_totals[category]['debit'] += data['debit']
        category_totals[category]['credit'] += data['credit']
        category_totals[category]['count'] += data['count']

        # Write data row
        ws.cell(row=row_idx, column=1, value=category)
        ws.cell(row=row_idx, column=2, value=subcategory)
        cell = ws.cell(row=row_idx, column=3, value=data['debit'])
        cell.number_format = CURRENCY_FORMAT
        cell = ws.cell(row=row_idx, column=4, value=data['credit'])
        cell.number_format = CURRENCY_FORMAT
        cell = ws.cell(row=row_idx, column=5, value=data['credit'] - data['debit'])
        cell.number_format = CURRENCY_FORMAT
        ws.cell(row=row_idx, column=6, value=data['count'])

        row_idx += 1

    # Last category subtotal
    if current_category:
        _write_category_subtotal(ws, row_idx, current_category, category_totals[current_category])
        row_idx += 1

    # Grand total
    row_idx += 1
    grand_total = {
        'debit': sum(d['debit'] for d in category_totals.values()),
        'credit': sum(d['credit'] for d in category_totals.values()),
        'count': sum(d['count'] for d in category_totals.values()),
    }
    ws.cell(row=row_idx, column=1, value="GRAND TOTAL").font = Font(bold=True)
    cell = ws.cell(row=row_idx, column=3, value=grand_total['debit'])
    cell.number_format = CURRENCY_FORMAT
    cell.font = Font(bold=True)
    cell = ws.cell(row=row_idx, column=4, value=grand_total['credit'])
    cell.number_format = CURRENCY_FORMAT
    cell.font = Font(bold=True)
    cell = ws.cell(row=row_idx, column=5, value=grand_total['credit'] - grand_total['debit'])
    cell.number_format = CURRENCY_FORMAT
    cell.font = Font(bold=True)
    ws.cell(row=row_idx, column=6, value=grand_total['count']).font = Font(bold=True)

    # Column widths
    column_widths = [20, 25, 18, 18, 18, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[_get_column_letter(col)].width = width

    ws.freeze_panes = "A2"


def _write_category_subtotal(ws: Worksheet, row: int, category: str, data: Dict) -> None:
    """Write a category subtotal row."""
    ws.cell(row=row, column=1, value=f"{category} Subtotal").font = Font(bold=True, italic=True)
    cell = ws.cell(row=row, column=3, value=data['debit'])
    cell.number_format = CURRENCY_FORMAT
    cell.font = Font(bold=True, italic=True)
    cell = ws.cell(row=row, column=4, value=data['credit'])
    cell.number_format = CURRENCY_FORMAT
    cell.font = Font(bold=True, italic=True)
    cell = ws.cell(row=row, column=5, value=data['credit'] - data['debit'])
    cell.number_format = CURRENCY_FORMAT
    cell.font = Font(bold=True, italic=True)
    ws.cell(row=row, column=6, value=data['count']).font = Font(bold=True, italic=True)


def _create_monthly_summary_sheet(wb: Workbook, transactions: List[Transaction]) -> None:
    """Create the Monthly Summary sheet."""
    ws = wb.create_sheet("Monthly Summary")

    # Aggregate by month
    monthly: Dict[str, Dict[str, float]] = defaultdict(
        lambda: {'debit': 0.0, 'credit': 0.0}
    )

    for txn in transactions:
        if txn.date:
            month_key = txn.date.strftime("%Y-%m")
            monthly[month_key]['debit'] += txn.debit or 0
            monthly[month_key]['credit'] += txn.credit or 0

    # Headers
    headers = ["Month", "Total Debits", "Total Credits", "Net Flow"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Sort by month
    sorted_months = sorted(monthly.keys())

    total_debit = 0.0
    total_credit = 0.0

    for row_idx, month in enumerate(sorted_months, 2):
        data = monthly[month]
        total_debit += data['debit']
        total_credit += data['credit']

        ws.cell(row=row_idx, column=1, value=month)
        cell = ws.cell(row=row_idx, column=2, value=data['debit'])
        cell.number_format = CURRENCY_FORMAT
        cell = ws.cell(row=row_idx, column=3, value=data['credit'])
        cell.number_format = CURRENCY_FORMAT
        cell = ws.cell(row=row_idx, column=4, value=data['credit'] - data['debit'])
        cell.number_format = CURRENCY_FORMAT

        if row_idx % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).fill = ALT_ROW_FILL

    # Total row
    row_idx = len(sorted_months) + 2
    ws.cell(row=row_idx, column=1, value="TOTAL").font = Font(bold=True)
    cell = ws.cell(row=row_idx, column=2, value=total_debit)
    cell.number_format = CURRENCY_FORMAT
    cell.font = Font(bold=True)
    cell = ws.cell(row=row_idx, column=3, value=total_credit)
    cell.number_format = CURRENCY_FORMAT
    cell.font = Font(bold=True)
    cell = ws.cell(row=row_idx, column=4, value=total_credit - total_debit)
    cell.number_format = CURRENCY_FORMAT
    cell.font = Font(bold=True)

    # Column widths
    for col, width in enumerate([15, 18, 18, 18], 1):
        ws.column_dimensions[_get_column_letter(col)].width = width

    ws.freeze_panes = "A2"


def _create_flagged_sheet(wb: Workbook, transactions: List[Transaction]) -> None:
    """Create the Flagged for Review sheet."""
    ws = wb.create_sheet("Flagged for Review")

    # Filter flagged transactions
    flagged = [t for t in transactions if t.categorization_source == "flagged"]

    # Headers
    headers = [
        "Date", "Description", "Debit", "Credit", "Balance",
        "AI Suggestion", "Your Category", "Your Subcategory"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Write data
    for row_idx, txn in enumerate(flagged, 2):
        cell = ws.cell(row=row_idx, column=1, value=txn.date)
        cell.number_format = DATE_FORMAT

        ws.cell(row=row_idx, column=2, value=txn.description)

        cell = ws.cell(row=row_idx, column=3, value=txn.debit)
        if txn.debit:
            cell.number_format = CURRENCY_FORMAT

        cell = ws.cell(row=row_idx, column=4, value=txn.credit)
        if txn.credit:
            cell.number_format = CURRENCY_FORMAT

        cell = ws.cell(row=row_idx, column=5, value=txn.balance)
        if txn.balance:
            cell.number_format = CURRENCY_FORMAT

        ws.cell(row=row_idx, column=6, value=txn.haiku_suggestion)

        # Empty columns for user to fill
        ws.cell(row=row_idx, column=7, value="")
        ws.cell(row=row_idx, column=8, value="")

        # Yellow background for flagged
        for col in range(1, 9):
            ws.cell(row=row_idx, column=col).fill = FLAGGED_FILL

    # Column widths
    for col, width in enumerate([12, 50, 15, 15, 15, 40, 20, 25], 1):
        ws.column_dimensions[_get_column_letter(col)].width = width

    # Add note at bottom
    if flagged:
        note_row = len(flagged) + 3
        ws.cell(row=note_row, column=1,
                value="Please fill in 'Your Category' and 'Your Subcategory' columns for flagged transactions.")
        ws.cell(row=note_row, column=1).font = Font(italic=True)

    ws.freeze_panes = "A2"


def _create_statistics_sheet(wb: Workbook, transactions: List[Transaction]) -> None:
    """Create the Statistics sheet."""
    ws = wb.create_sheet("Statistics")

    # Calculate statistics
    total = len(transactions)
    rules_count = sum(1 for t in transactions if t.categorization_source == "rules")
    haiku_count = sum(1 for t in transactions if t.categorization_source == "haiku")
    flagged_count = sum(1 for t in transactions if t.categorization_source == "flagged")

    dates = [t.date for t in transactions if t.date]
    date_range = f"{min(dates)} to {max(dates)}" if dates else "N/A"

    total_debit = sum(t.debit or 0 for t in transactions)
    total_credit = sum(t.credit or 0 for t in transactions)

    # Category counts
    category_counts: Dict[str, int] = defaultdict(int)
    category_amounts: Dict[str, float] = defaultdict(float)
    for t in transactions:
        if t.category:
            category_counts[t.category] += 1
            category_amounts[t.category] += abs(t.debit or 0) + abs(t.credit or 0)

    # Write statistics
    stats = [
        ("Summary Statistics", ""),
        ("", ""),
        ("Total Transactions", total),
        ("Date Range", date_range),
        ("Total Debits", total_debit),
        ("Total Credits", total_credit),
        ("Net Flow", total_credit - total_debit),
        ("", ""),
        ("Categorization Breakdown", ""),
        ("Rules Matched", f"{rules_count} ({rules_count/total*100:.1f}%)" if total else "0"),
        ("Haiku Matched", f"{haiku_count} ({haiku_count/total*100:.1f}%)" if total else "0"),
        ("Flagged for Review", f"{flagged_count} ({flagged_count/total*100:.1f}%)" if total else "0"),
    ]

    row_idx = 1
    for label, value in stats:
        cell = ws.cell(row=row_idx, column=1, value=label)
        if label and not value:
            cell.font = Font(bold=True, size=12)
        if isinstance(value, float):
            cell = ws.cell(row=row_idx, column=2, value=value)
            cell.number_format = CURRENCY_FORMAT
        else:
            ws.cell(row=row_idx, column=2, value=value)
        row_idx += 1

    # Top categories by count
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Top 10 Categories by Count").font = Font(bold=True, size=12)
    row_idx += 1

    top_by_count = sorted(category_counts.items(), key=lambda x: x[1], reverse=True)[:10]
    for cat, count in top_by_count:
        ws.cell(row=row_idx, column=1, value=cat)
        ws.cell(row=row_idx, column=2, value=count)
        row_idx += 1

    # Top categories by amount
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Top 10 Categories by Amount").font = Font(bold=True, size=12)
    row_idx += 1

    top_by_amount = sorted(category_amounts.items(), key=lambda x: x[1], reverse=True)[:10]
    for cat, amount in top_by_amount:
        ws.cell(row=row_idx, column=1, value=cat)
        cell = ws.cell(row=row_idx, column=2, value=amount)
        cell.number_format = CURRENCY_FORMAT
        row_idx += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25


def _get_column_letter(col_num: int) -> str:
    """Convert column number to letter (1 = A, 27 = AA, etc.)."""
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result
